# apps/gso_reports/views.py
import os
import csv
import json
import calendar
import openpyxl
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt

from apps.gso_requests.models import ServiceRequest, Feedback
from apps.gso_accounts.models import User, Unit
from .models import WorkAccomplishmentReport, SuccessIndicator, IPMT
from .utils import normalize_report
from apps.ai_service.utils import generate_war_description


# -------------------------------
# Role Checks
# -------------------------------
def is_gso_or_director(user):
    return user.is_authenticated and user.role in ["gso", "director"]


# -------------------------------
# Accomplishment Report View
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def accomplishment_report(request):
    completed_requests = ServiceRequest.objects.filter(status="Completed").order_by("-created_at")
    all_wars = WorkAccomplishmentReport.objects.select_related("request", "unit") \
        .prefetch_related("assigned_personnel").all().order_by("-date_started")

    reports = []
    war_request_ids = set(war.request_id for war in all_wars if war.request_id)

    # Process completed requests (no WAR yet)
    for r in completed_requests:
        if r.id in war_request_ids:
            continue
        norm = normalize_report(r)
        norm["id"] = r.id

        if not norm.get("description") or not norm["description"].strip():
            try:
                desc = generate_war_description(
                    activity_name=getattr(r, "activity_name", getattr(r, "title", "Task")),
                    unit=getattr(r.unit, "name", None),
                    personnel_names=[p.get_full_name() for p in r.assigned_personnel.all()] if hasattr(r, "assigned_personnel") else None
                )
                r.description = desc or "No description generated."
                r.save(update_fields=["description"])
                norm["description"] = r.description
            except Exception as e:
                norm["description"] = f"Error generating description: {e}"

        reports.append(norm)

    # Process existing WARs
    for war in all_wars:
        norm = normalize_report(war)
        norm["id"] = war.id

        if not norm.get("description") or not norm["description"].strip():
            try:
                desc = generate_war_description(
                    activity_name=getattr(war, "activity_name", getattr(war, "title", "Task")),
                    unit=getattr(war.unit, "name", None),
                    personnel_names=[p.get_full_name() for p in war.assigned_personnel.all()] if hasattr(war, "assigned_personnel") else None
                )
                war.description = desc or "No description generated."
                war.save(update_fields=["description"])
                norm["description"] = war.description
            except Exception as e:
                norm["description"] = f"Error generating description: {e}"

        reports.append(norm)

    # Filters
    search_query = request.GET.get("q")
    if search_query:
        reports = [r for r in reports if search_query.lower() in str(r).lower()]

    unit_filter = request.GET.get("unit")
    if unit_filter:
        reports = [r for r in reports if r["unit"].lower() == unit_filter.lower()]

    reports.sort(key=lambda r: r["date"], reverse=True)

    personnel_qs = User.objects.filter(role="personnel", account_status="active") \
        .select_related('unit').order_by('unit__name', 'first_name')

    personnel_list = [
        {
            "full_name": u.get_full_name() or u.username,
            "username": u.username,
            "unit": u.unit.name.lower() if u.unit else "unassigned"
        }
        for u in personnel_qs
    ]

    return render(
        request,
        "gso_office/accomplishment_report/accomplishment_report.html",
        {"reports": reports, "personnel_list": personnel_list},
    )

# -------------------------------
# 10/28/25 not edited below or walang kwenta
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
@csrf_exempt  # allows AJAX POST from JS
def update_success_indicator(request):
    """
    Updates the Success Indicator of a Work Accomplishment Report via AJAX.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            war_id = data.get("war_id")
            indicator_id = data.get("indicator_id")

            war = WorkAccomplishmentReport.objects.get(id=war_id)
            indicator = SuccessIndicator.objects.get(id=indicator_id)

            war.success_indicator = indicator
            war.save(update_fields=["success_indicator"])

            return JsonResponse({"success": True, "indicator": indicator.code})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request"})
# -------------------------------
# Generate IPMT Excel
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def generate_ipmt(request):
    reports = []
    personnel_list = []

    if request.method == "POST":
        try:
            body = json.loads(request.body.decode("utf-8"))
            month_filter = body.get("month")
            unit_filter = body.get("unit")
            personnel_param = body.get("personnel", "")
            reports = body.get("rows", [])
        except Exception:
            month_filter = request.POST.get("month")
            unit_filter = request.POST.get("unit")
            personnel_param = request.POST.get("personnel", "")
            rows_data = request.POST.get("rows", "[]")
            try:
                reports = json.loads(rows_data)
            except json.JSONDecodeError:
                reports = []

        personnel_list = [p.strip() for p in personnel_param.split(",") if p.strip()]

        for r in reports:
            if not r.get("indicator"):
                continue
            code_only = r["indicator"].split(" - ")[0].strip()
            si = SuccessIndicator.objects.filter(code__iexact=code_only).first()
            if si:
                r["indicator"] = f"{si.code} - {si.description}"

    else:
        return HttpResponse("Only POST allowed.", status=400)

    template_path = os.path.join(settings.BASE_DIR, "static", "excel_file", "sampleipmt.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    personnel_fullnames = []
    for identifier in personnel_list:
        user_obj = get_user_by_identifier(identifier)
        if user_obj:
            full_name = (user_obj.get_full_name() or "").strip()
            personnel_fullnames.append(full_name or user_obj.username)
        else:
            personnel_fullnames.append(identifier)

    ws["B8"] = ", ".join(personnel_fullnames) if personnel_fullnames else "No personnel found"

    if "-" in month_filter:
        year, month_num = map(int, month_filter.split("-"))
        month_name = f"{calendar.month_name[month_num]} {year}"
    else:
        month_name = month_filter
    ws["B11"] = month_name

    start_row = 13
    for i, r in enumerate(reports, start=start_row):
        ws.cell(row=i, column=1).value = r.get("indicator", "")
        ws.cell(row=i, column=2).value = r.get("description", "")
        ws.cell(row=i, column=3).value = r.get("remarks", "")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"IPMT_{unit_filter}_{month_filter}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------------------------------
# Helper
# -------------------------------
def get_user_by_identifier(identifier):
    identifier = identifier.strip()
    if not identifier:
        return None

    user = User.objects.filter(username__iexact=identifier).first()
    if user:
        return user

    parts = identifier.split()
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        user = User.objects.filter(Q(first_name__iexact=first) & Q(last_name__iexact=last)).first()
        if user:
            return user

    return User.objects.filter(Q(first_name__icontains=identifier) | Q(last_name__icontains=identifier)).first()

# -------------------------------
# Get WAR Description (AJAX)
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def get_war_description(request, war_id):
    try:
        war = WorkAccomplishmentReport.objects.get(id=war_id)
        return JsonResponse({'description': war.description or ""})
    except WorkAccomplishmentReport.DoesNotExist:
        return JsonResponse({'error': 'WAR not found'}, status=404)
    

# -------------------------------
# Preview IPMT (Web)
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def preview_ipmt(request):
    """
    Preview IPMT rows for the selected unit, personnel, and month.
    Fetches WARs linked to each SuccessIndicator and assigned personnel.
    """
    month_filter = request.GET.get("month")
    unit_filter = request.GET.get("unit")
    personnel_names = request.GET.getlist("personnel[]") or []

    if not month_filter:
        return HttpResponse("Month is required in 'YYYY-MM' format.", status=400)

    try:
        year, month_num = map(int, month_filter.split("-"))
    except ValueError:
        return HttpResponse("Invalid month format. Use YYYY-MM.", status=400)

    unit = Unit.objects.filter(name__iexact=unit_filter).first()
    if not unit:
        return HttpResponse("Unit not found.", status=404)

    reports = []

    for person_name in personnel_names:
        user = get_user_by_identifier(person_name)
        if not user:
            continue

        # Fetch all active success indicators under the unit
        indicators = SuccessIndicator.objects.filter(unit=unit, is_active=True)

        for indicator in indicators:
            # Get WARs for this user and indicator within the selected month
            wars = WorkAccomplishmentReport.objects.filter(
                unit=unit,
                assigned_personnel=user,
                success_indicator=indicator,
                date_started__year=year,
                date_started__month=month_num
            )

            # Combine descriptions from all WARs
            description = " ".join([w.description for w in wars if w.description]) or ""

            reports.append({
                "indicator": indicator.code,
                "description": description,
                "remarks": "COMPLIED" if description else "",
                "war_ids": [w.id for w in wars],
            })

    context = {
        "reports": reports,
        "month_filter": month_filter,
        "unit_filter": unit_filter,
        "personnel_names": personnel_names,
    }

    return render(request, "gso_office/ipmt/ipmt_preview.html", context)

# -------------------------------
# Save IPMT
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def save_ipmt(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)

    try:
        data = json.loads(request.body)
        month = data.get("month")
        unit_name = data.get("unit")
        personnel_names = data.get("personnel", [])
        rows = data.get("rows", [])
    except Exception as e:
        return JsonResponse({"error": f"Invalid JSON: {str(e)}"}, status=400)

    unit = Unit.objects.filter(name__iexact=unit_name).first()
    if not unit:
        return JsonResponse({"error": "Unit not found"}, status=404)

    for person_name in personnel_names:
        user = get_user_by_identifier(person_name)
        if not user:
            continue

        for row in rows:
            indicator_code = row.get("indicator")
            if not indicator_code:
                continue

            indicator = SuccessIndicator.objects.filter(unit=unit, code__iexact=indicator_code).first()
            if not indicator:
                # Create missing indicator automatically
                indicator = SuccessIndicator.objects.create(
                    unit=unit,
                    code=indicator_code,
                    description=row.get("description", ""),
                    is_active=True
                )

            # WARs linked to this indicator
            war_ids = row.get("war_ids", [])
            wars = WorkAccomplishmentReport.objects.filter(
                assigned_personnel=user,
                unit=unit
            )
            if war_ids:
                wars = wars.filter(id__in=war_ids)

            accomplishment = row.get("description", "").strip()
            remarks = row.get("remarks", "").strip() or accomplishment

            # Save or update IPMT
            ipmt_obj, _ = IPMT.objects.update_or_create(
                personnel=user,
                unit=unit,
                month=month,
                indicator=indicator,
                defaults={
                    "accomplishment": accomplishment,
                    "remarks": remarks
                }
            )

            # Link WARs to IPMT
            ipmt_obj.reports.set(wars)

    return JsonResponse({"status": "success"})



# -------------------------------
# GSO OFFICE: Edit Success Indicator in WAR
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def update_war_success_indicator(request, war_id):
    """
    Allow GSO Office to edit the Success Indicator directly from the WAR page.
    """
    war = get_object_or_404(WorkAccomplishmentReport, id=war_id)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()

        if name:
            indicator, _ = SuccessIndicator.objects.get_or_create(name=name)
            if description:
                indicator.description = description
                indicator.save(update_fields=["description"])
            war.success_indicator = indicator
            war.save(update_fields=["success_indicator"])
            messages.success(request, "Success Indicator updated for WAR.")
            return redirect("gso_reports:war_detail", war_id=war.id)

        messages.warning(request, "Indicator name is required.")

    return render(request, "gso_office/war_success_indicator_form.html", {
        "war": war,
    })














# GSO Analytics View
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from apps.gso_requests.models import ServiceRequest as Request
from apps.gso_inventory.models import InventoryItem as Material

@login_required
def gso_analytics(request):
    # ===== REQUEST ANALYTICS =====
    total_requests = Request.objects.count()
    completed_requests = Request.objects.filter(status='Completed').count()
    pending_requests = Request.objects.filter(status='Pending').count()
    in_progress_requests = Request.objects.filter(status='In Progress').count()

    # ===== INVENTORY ANALYTICS =====
    total_materials = Material.objects.count()
    low_stock_materials = Material.objects.filter(quantity__lte=10).count()  # threshold for low stock
    out_of_stock = Material.objects.filter(quantity=0).count()

    # ===== CONTEXT =====
    context = {
        # Request analytics
        'total_requests': total_requests,
        'completed_requests': completed_requests,
        'pending_requests': pending_requests,
        'in_progress_requests': in_progress_requests,

        # Inventory analytics
        'total_materials': total_materials,
        'low_stock_materials': low_stock_materials,
        'out_of_stock': out_of_stock,
    }

    return render(request, 'gso_office/analytics/gso_analytics.html', context)


@login_required
@user_passes_test(is_gso_or_director)
def feedback_reports(request):
    """Show and export all feedback records (using requestor info only)."""
    feedback_list = Feedback.objects.select_related("request", "request__requestor").order_by("-date_submitted")

    # ✅ Export to CSV
    if "export" in request.GET:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="feedback_report.csv"'

        writer = csv.writer(response)
        writer.writerow([
            "Service Request ID",
            "Requestor Name",
            "Requestor Email",
            "SQD1", "SQD2", "SQD3", "SQD4", "SQD5", "SQD6", "SQD7", "SQD8", "SQD9",
            "CC1", "CC2", "CC3",
            "Average Score",
            "Suggestions",
            "Date Submitted"
        ])

        for fb in feedback_list:
            req = fb.request

            # ✅ Get requestor info (custom > user)
            if req:
                requestor_name = req.custom_full_name or (req.requestor.get_full_name() if req.requestor else "")
                requestor_email = req.custom_email or (req.requestor.email if req.requestor else "")
                request_id = req.id
            else:
                requestor_name = ""
                requestor_email = ""
                request_id = ""

            # ✅ Format date properly
            formatted_date = fb.date_submitted.strftime("%Y-%m-%d %H:%M") if fb.date_submitted else ""

            writer.writerow([
                request_id,
                requestor_name,
                requestor_email,
                fb.sqd1 or "", fb.sqd2 or "", fb.sqd3 or "", fb.sqd4 or "",
                fb.sqd5 or "", fb.sqd6 or "", fb.sqd7 or "", fb.sqd8 or "", fb.sqd9 or "",
                fb.cc1 or "", fb.cc2 or "", fb.cc3 or "",
                round(fb.average_score, 2),
                fb.suggestions or "",
                formatted_date
            ])

        return response

    # ✅ Compute average for HTML table view
    for fb in feedback_list:
        scores = [fb.sqd1, fb.sqd2, fb.sqd3, fb.sqd4, fb.sqd5, fb.sqd6, fb.sqd7, fb.sqd8, fb.sqd9]
        valid_scores = [s for s in scores if s is not None]
        fb.average_rating = round(sum(valid_scores) / len(valid_scores), 2) if valid_scores else 0

    return render(
        request,
        "gso_office/feedbacks/feedback_reports.html",
        {"feedback_list": feedback_list}
    )
